# Utility functions are stored in here ofr clearer code structure
import numpy
import openpyxl
import xlrd
import pandas
from multiprocessing import Queue
import multiprocessing
from os import walk
from os.path import exists
import time
from datetime import datetime




def write_to_csv(path, csv_row):  # Write output files from Source_2
    with open(path, 'a', newline='', encoding="utf-8") as f:
        f.write(csv_row)

def csv2queue(file_path, q):
    try:
        print(file_path)
        dataList = pandas.read_csv(file_path, index_col=False, header=None)
        data_arr = numpy.array(dataList.values)
        for i in range(len(data_arr)):
            q.put(data_arr[i][0])
    except Exception as e:
        print(e)

# Converts an xlsx file to csv
def xlsx2csv(input_file, output_file):
    try:
        output_file = output_file +  + "/output.csv"
        if(input_file[-5:] == '.xlsx'):
            wb_obj = openpyxl.load_workbook(input_file)
            sheet = wb_obj.active
            for row in sheet.iter_rows(max_row=sheet.max_row):
                temp = ""
                row_curr = "\""
                cell = 0
                for cell in range(sheet.max_column):
                    temp = str(row[cell].value)
                    temp = temp.replace("\"", "\"\"")
                    row_curr += temp + "\",\""
                    cell +=1
                write_to_csv(output_file, (row_curr[:-2] + "\n"))
        else:
            raise Exception("File extension is wrong")
    except Exception as e:
        print(e)

# Takes file1 csv, takes file2 csv
# Compare two files by their given column indexes (file1_p, file2_p)
# Creates an output file according to the difference between two files from file1
def compare2diff(input1, input2, output):
    try:
        q = Queue()
        opt = [3, 4, 5]
        create_queue_infile(input1, q, opt)
        pool = multiprocessing.Pool(processes=(multiprocessing.cpu_count() - 1))
        while not (q.empty()):
            res = pool.apply_async(substract_from_file, args=(q.get(), input2, 5, output,))
        pool.close()
        pool.join()
        q.close()
    except Exception as e:
        print(e)

# Creates a queue from the rows of a given csv file
# path -> path to csv file, q -> queue
def create_queue_infile_raw(path, q):
    try:
        data_list = pandas.read_csv(path, index_col=False, header=None)
        data_arr = numpy.array(data_list.values)
        for i in range(len(data_arr)):
            q.put(str(data_arr[i][0]))
    except Exception as e:
        print(e)

def create_queue_infile(path, q, opt):
    try:
        value = ""
        data_list = pandas.read_csv(path, index_col=False, header=None)
        data_arr = numpy.array(data_list.values)
        for i in range(len(data_arr)):
            for j in opt:
                value = value + str(data_arr[i][j]) + " "
            value = value[:-1]
            q.put(value)
            value = ""
    except Exception as e:
        print(e)


def create_files_queue_from_directory(path, q):
    try:
        filenames_s = next(walk(path), (None, None, []))[2]
        for file in filenames_s:
            file_path = path + '/' + file
            q.put(file_path, True, None)
    except Exception as e:
        print(e)

# Search for a given field in a given file with given column index
# Write to another file if not found
def substract_from_file(field, file, col_index, output):
    try:
        output = output + "/output_sub.csv"
        found = False
        data_list = pandas.read_csv(file, low_memory=False, index_col=False, header=None)
        data_arr = numpy.array(data_list.values)
        for i in range(len(data_arr)):
            if ( transliterate_to_en(str(field).upper()) in transliterate_to_en(str(data_arr[i][col_index]).upper()) ) or ( transliterate_to_en_v2(str(field).upper()) in transliterate_to_en_v2(str(data_arr[i][col_index]).upper()) ):
                if str(field).upper() != "NAN NAN NAN":
                    found = True
        if not found:
            result = str(field + "\n").upper()
            write_to_csv(output, result)

    except Exception as e:
        print(e)


# Filters a file according to the values in the base file
# file1 -> Base File, file2 -> Data File, output -> Output file location to create an intersection list
def intersection_of_file(file1, file2, output):
    try:
        output = output + "/output_int.csv"
        q = Queue()
        create_queue_infile_raw(file1, q)
        pool = multiprocessing.Pool(processes=(multiprocessing.cpu_count() - 1))
        while not (q.empty()):
            res = pool.apply_async(create_intersection, args=(q.get(), file2, output, ))
        pool.close()
        pool.join()
        q.close()
    except Exception as e:
        print(e)


# Creates an intersection list
# Takes an input row and another list to check if it exist there to get information
# Writes the result to a certain file
def create_intersection(input, file, output):
    try:
        data_list = pandas.read_csv(file, low_memory=False, index_col=False, header=None)
        data_arr = numpy.array(data_list.values)
        for i in range(len(data_arr)):
            string = str(data_arr[i][3]) + " " + str(data_arr[i][4]) + " " + str(data_arr[i][5])
            if str(input).upper() == str(string).upper() :
                row = ""
                for j in range(len(data_arr[i])):
                    row = row + "\"" + str(data_arr[i][j]) + "\","
                row = row[:-2] + "\"\n"
                write_to_csv(output, row)
    except Exception as e:
        print(e)


# Transliterate strings from Azerbaijani Latin to English Latin
# Version 2 differs from Version 1 by converting 'ə' to 'e' or 'a'
def transliterate_to_en (string):
    ret = string
    ret = ret.replace('Ç','C')
    ret = ret.replace('ç','c')
    ret = ret.replace('İ','I')
    ret = ret.replace('i','i')
    ret = ret.replace('I','I')
    ret = ret.replace('ı','i')
    ret = ret.replace('Ğ','G')
    ret = ret.replace('ğ','g')
    ret = ret.replace('Ö','O')
    ret = ret.replace('ö','o')
    ret = ret.replace('Ş','S')
    ret = ret.replace('ş','s')
    ret = ret.replace('Ü','U')
    ret = ret.replace('ü','u')
    ret = ret.replace('Ə','E')
    ret = ret.replace('ə','e')
    return str(ret).upper()


def transliterate_to_en_v2 (string):
    ret = string
    ret = ret.replace('Ç','C')
    ret = ret.replace('ç','c')
    ret = ret.replace('İ','I')
    ret = ret.replace('i','i')
    ret = ret.replace('I','I')
    ret = ret.replace('ı','i')
    ret = ret.replace('Ğ','G')
    ret = ret.replace('ğ','g')
    ret = ret.replace('Ö','O')
    ret = ret.replace('ö','o')
    ret = ret.replace('Ş','S')
    ret = ret.replace('ş','s')
    ret = ret.replace('Ü','U')
    ret = ret.replace('ü','u')
    ret = ret.replace('Ə','A')
    ret = ret.replace('ə','a')
    return str(ret).upper()


def create_date_list_master(dir_path, output):
    print(dir_path)
    q = Queue()
    create_files_queue_from_directory(dir_path,q)
    pool = multiprocessing.Pool(processes=(multiprocessing.cpu_count() - 1))
    while not (q.empty()):
        res = pool.apply_async(create_date_dist_worker, args=(q.get(), output,))
    pool.close()
    pool.join()
    q.close()


# Create distinct files based on date from the external data source
def create_date_dist_worker(file_path, output):
    try:
        wb_obj = openpyxl.load_workbook(file_path)
        sheet = wb_obj.active
        date = ""
        row_curr = ""
        for row in sheet.iter_rows(max_row=sheet.max_row):
            if row[1].value != "Tarix":
                date = str(row[1].value)
                date = date.split(' ')[0]
                date = datetime.strptime(date, '%Y-%m-%d').strftime('%Y%m%d')
                row_curr = "\""+str(date) +"\",\""+ str(row[3].value) +"\",\""+ str(row[4].value) +"\",\""+ str(row[6].value) +"\",\""+ str(row[8].value) + "\"\n"
                print(row_curr)
                write_to_csv(output+'/'+date+'.csv', row_curr)
                date = ""
                row_curr = ""
    except Exception as e:
        print(e)


def cross_ref_org2corr_worker(corr_file_path, original_folder_path, output_folder_path):
    org = Queue()
    create_files_queue_from_directory(original_folder_path, org)
    print(corr_file_path)
    rowIndex = 0
    try:
        wb_obj = xlrd.open_workbook(corr_file_path)
        sheet = wb_obj.sheet_by_index(0)
        while rowIndex in range(sheet.nrows):
            flag = False
            # Compare 0, 4, 5 from corrupted with 1, 6, 8 from original
            while not (org.empty()):
                wb_org = openpyxl.load_workbook(org.get())
                sheet_org = wb_org.active
                for row in sheet_org.iter_rows(max_row=sheet_org.max_row):
                    try:

                        if (datetime.strptime(str(sheet.cell_value(rowIndex, 0)), "%d-%m-%Y").date() == datetime.date(row[1].value)):
                            #print(datetime.strptime(str(sheet.cell_value(rowIndex, 0)), "%d-%m-%Y").date(),
                            #      datetime.date(row[1].value))
                            if (sheet.cell_value(rowIndex, 4) == row[6].value):
                                if (sheet.cell_value(rowIndex, 5) == row[8].value):
                                    flag = True
                    except Exception as ex:
                        continue
            if not flag:
                problem_row = str(sheet.cell_value(rowIndex, 0)) + "," + str(sheet.cell_value(rowIndex, 4)) + "," + str(sheet.cell_value(rowIndex, 5)) + "\n"
                output_file_path = output_folder_path + "/" + str(sheet.cell_value(rowIndex, 0)) + ".csv"
                write_to_csv(output_file_path, problem_row)
    except Exception as e:
        print(e)


def cross_ref_org2corr(original_folder_path, corrupted_folder_path, output_folder_path):
    corr = Queue()
    create_files_queue_from_directory(corrupted_folder_path, corr)
    pool = multiprocessing.Pool(processes=(multiprocessing.cpu_count() - 1))
    while not (corr.empty()):
        res = pool.apply_async(cross_ref_org2corr_worker, args=(corr.get(), original_folder_path, output_folder_path,))
    pool.close()
    pool.join()
    corr.close()