import csv
import multiprocessing
import string
from multiprocessing import Queue
from os import walk
from os.path import exists
import openpyxl as openpyxl
from datetime import datetime
from tkinter import *
from tkinter import filedialog
import pandas
import numpy
import time
import xlrd
#custom imports
import utility

# Performance
# Generate new Invoice list from system data 57s
# Generate new Invoice list and iterate through info list 833s
# Generate new Invoice list with the information gathered from info list 1048s
# Validate data 114s
# Generate new set of files based on system source 1837s
# Remove duplicates in generated files 3.9s
# Convert xlsx to csv (38k*16 sample size)
#

output_folder = 'SingleFileOutput/'  # Folder to write output
data_folder = 'SingleFileInput/'   # Folder that contains data to mine
infolist_folder = 'Invoice/stud_based_invoice_list/clean_lvl2/' # Folder that contains reference info
input_file = r"C:\Users\yavuz\PycharmProjects\Finance_Mine\DATAMINING\SingleFileInput\students.csv" # Input file path for single file operations
input_file2 = r"C:\Users\yavuz\PycharmProjects\Finance_Mine\DATAMINING\SingleFileInput2\Təhsilalanlar_ATIS.csv"
output_file = r"C:\Users\yavuz\PycharmProjects\Finance_Mine\DATAMINING\SingleFileOutput\diff_stud.csv" # Output file path for single file operations


# Out -> Invoice/stud_based_invoice_list/clean_lvl3/
# In -> Invoice/stud_based_invoice_list/clean_lvl1/ & Invoice/stud_based_invoice_list/clean_lvl2/
def check_base_info(data, file_list):
    path = ["Veriler1_old/E_SPREADSHEET/V_butun kurslar.csv","Veriler1_old/E_SPREADSHEET_HARICI/V_butun kurslar.csv"]
    data_csv = pandas.read_csv(data_folder+data, index_col=False, header=None)
    data_csv_arr = numpy.array(data_csv.values)
    output_path = output_folder + data
    info_total = int(0)
    data_total = int(0)
    flag = int(0)
    try:
        for path_in in path:
            info_csv = pandas.read_csv(path_in)
            info_csv_arr = numpy.array(info_csv.values)
            for i in range(len(info_csv_arr)):
                if str(data[:-4]) == str(info_csv_arr[i][0]):
                    # print(str(info_csv_arr[i][9]) +" "+ str(info_csv_arr[i][10]) +" "+ str(info_csv_arr[i][11]) +" "+ str(info_csv_arr[i][12]) +" "+ str(info_csv_arr[i][13]))
                    info_total = int(info_csv_arr[i][9]) + int(info_csv_arr[i][10]) + int(info_csv_arr[i][11]) + int(info_csv_arr[i][12]) + int(info_csv_arr[i][13])
                    for j in range(len(data_csv_arr)):
                        data_total = data_total + int(data_csv_arr[j][2])

                    if(data_total == info_total) or (data not in file_list):
                        flag = 1
        if flag == 0:
            data_csv.to_csv(output_path, index=False, header=None, quoting=csv.QUOTE_ALL)
    except Exception as e:
        print(e)

# Out -> clean_lvl2
# In -> clean_lvl1
def cross_check_info_data(data_file):
    try:
        data_list = list()
        data_csv = pandas.read_csv(data_folder + data_file, index_col=False, header=None)
        data_csv_arr = numpy.array(data_csv.values)
        for i in range(len(data_csv_arr)):
            flag = int(0)
            info_path = infolist_folder + str(data_csv_arr[i][1]) + ".csv"
            info_csv = pandas.read_csv(info_path, index_col=False, header=None)
            info_csv_arr = numpy.array(info_csv.values)
            names = str(data_csv_arr[i][3]).split()
            for j in range(len(info_csv_arr)):
                check_count = float(0)
                if numpy.round(float(info_csv_arr[j][3]),1) == numpy.round(float(data_csv_arr[i][2]),1):
                    for name in names:
                        if (utility.transliterate_to_en(str(name)) in utility.transliterate_to_en(str(info_csv_arr[j][4]))) or (utility.transliterate_to_en(str(name)) in utility.transliterate_to_en(str(info_csv_arr[j][1]))) or (utility.transliterate_to_en_v2(str(name)) in utility.transliterate_to_en_v2(str(info_csv_arr[j][4]))) or (utility.transliterate_to_en_v2(str(name)) in utility.transliterate_to_en_v2(str(info_csv_arr[j][1]))):
                            check_count = check_count + 1
                        # else:
                        #     if transliterate_to_en(str(name)) == 'ROVSEN':
                        #         print(transliterate_to_en(str(name)) + " " + transliterate_to_en(str(info_csv_arr[j][4])) +" "+ transliterate_to_en(str(info_csv_arr[j][1])) +" "+ transliterate_to_en_v2(str(info_csv_arr[j][4])) +" "+ transliterate_to_en_v2(str(info_csv_arr[j][1])) + "\n")

                            # print(str(name) + " - " + str(info_arr[j][4]) + " - " + str(percentage))
                    if((check_count*100/len(names)) >= 65):
                        flag = 1
            if flag == 0:
                data_list.append(data_csv_arr[i]) # TODO writes only one FIX IT!
        for row in data_list:
            curr_row = "\"" + str(row[0]) + "\",\"" + str(row[1]) + "\",\"" + str(row[2]) + "\",\"" + str(row[3]) + "\"\n"
            utility.write_to_csv(output_folder+data_file, curr_row)

    except Exception as e:
        print(e)

# Any csv
# Out -> clean_lvl1
# In -> raw
def remove_duplicates(file_name):
    try:
        file_path = data_folder + file_name
        output_path = output_folder + file_name
        csv_file = pandas.read_csv(file_path, low_memory=False, index_col=False, header=None)
        csv_file.drop_duplicates(subset=None, inplace=True)
        csv_file.to_csv(output_path, index=False, header=None, quoting=csv.QUOTE_ALL)
    except Exception as e:
        print(e)

# Out -> stud_based_invoice/raw
# In -> stud_list_from_sys & Invoice
def create_stud_based_files(file_name, date_list):
    out_write_path = output_folder + file_name + ".csv"
    try :
        for i in range(len(date_list)):
            full_path = infolist_folder + date_list[i] + ".csv"
            data_list_arr = pandas.read_csv(full_path)
            data_list_arr_val = numpy.array(data_list_arr.values)
            for j in range(len(data_list_arr_val)):
                if data_list_arr_val[j][2] == file_name:
                    row = "\"" + str(file_name) + "\",\"" + str(data_list_arr_val[j][0]) + "\",\"" + str(data_list_arr_val[j][1]) + "\",\"" + str(data_list_arr_val[j][3]) + "\"\n"
                    utility.write_to_csv(out_write_path, row)
    except Exception as e:
        print(e)


# Letter 'ə' is inconsistent since it can be used as A or E when transformed, therefore it is being skipped

def cross_ref_invoice(data_file):
    try:
        info_csv = pandas.read_csv(infolist_folder + data_file)
        data_csv = pandas.read_csv(data_folder + data_file)
        data_arr = numpy.array(data_csv.values)
        info_arr = numpy.array(info_csv.values)
        for i in range(len(data_arr)):
            for j in range(len(info_arr)):
                if info_arr[j][3] == data_arr[i][1]:
                    names = data_arr[i][3].split()
                    i = len(names)
                    percentage = float(0)
                    for name in names:
                        if utility.transliterate_to_en(str(name)) in utility.transliterate_to_en(str(info_arr[j][4])):
                            percentage = percentage + 1
                            # print(str(name) + " - " + str(info_arr[j][4]) + " - " + str(percentage))
                    if((percentage*100/i) >= 65):
                        txt = " - " + str(names) + str(info_arr[j][4])
                        # print(round((percentage*100)/i,5), transliterate_to_en(txt) )
                        # print(str(str((percentage*100/i)) + " - " + info_arr[j][4]) + " - " + str(data_arr[i][3]) + "\n")
                    # print(str(info_arr[j]) + " - " + str(data_arr[i]))

    except Exception as e:
        print(e)

def validate_data(arg1): # Create a list with summed up results for validation
    id = ""
    value = ""
    curr_row = ""
    l = 0
    path = data_folder + arg1
    out_path = output_folder + "/outList.csv"
    dataList = pandas.read_csv(path)
    data_arr = numpy.array(dataList.values)
    for row in range(len(data_arr)):
        l = 0
        try:
            id = data_arr[row][2]
            value = data_arr[row][3]
            try:
                if exists(out_path) :
                    infoList = pandas.read_csv(out_path, header=None)
                    info_arr = numpy.array(infoList.values)
                    for index in range(len(info_arr)):
                        if info_arr[index][0] == id: # Check for duplicates
                            l=1
                            index = len(info_arr)

            except:
                print("Error at: " + id)
                l=0

            if l == 0:
                curr_row = "\"" + id + "\",\"" + value + "\"\n"
                utility.write_to_csv(out_path, curr_row)
        except:
            l=0


def get_student_info (arg2): # arg1: path to student file, arg2: student id
    path = ["Veriler1_old/E_SPREADSHEET/V_1 kurs.csv","Veriler1_old/E_SPREADSHEET/V_2 kurs.csv","Veriler1_old/E_SPREADSHEET/V_3 kurs.csv","Veriler1_old/E_SPREADSHEET/V_4 kurs.csv","Veriler1_old/E_SPREADSHEET/V_5 kurs.csv","Veriler1_old/E_SPREADSHEET_HARICI/V_1 kurs.csv", "Veriler1_old/E_SPREADSHEET_HARICI/V_2 kurs.csv","Veriler1_old/E_SPREADSHEET_HARICI/V_3 kurs.csv","Veriler1_old/E_SPREADSHEET_HARICI/V_4 kurs.csv","Veriler1_old/E_SPREADSHEET_HARICI/V_5 kurs.csv"]
    stu_info = "NO_INFO"
    for file in range(len(path)):
        personList = pandas.read_csv(path[file])
        info_arr = numpy.array(personList.values)
        for row in range(len(info_arr)):
            try:
                if arg2 == info_arr[row][0]:
                    stu_info = info_arr[row][1]

            except:
                l=0

    return stu_info

# Out -> Output/outlist.csv -> stud_list_from_sys
# In -> Invoice/ & Veriler*/
def create_cross_ref(arg1, arg2): # arg1: dates, arg2: invoice files
    path_to_out = 'Invoice/Output/' + str(arg1)
    #print(path_to_out)
    for file in arg2:
        path_to_csv = 'Invoice/' + file + '.csv'
        csvFile = pandas.read_csv(path_to_csv)
        # csvFile['G_tarix'] = pandas.to_datetime(csvFile['G_tarix'], format='%Y%m%d', errors='coerce')
        arr = numpy.array(csvFile.values)
        for row in range(len(arr)):
            try :
                if isinstance(arr[row][2], float):
                    date1 = xlrd.xldate_as_datetime(arr[row][2],0)
                    date2 = date1.date()
                    date3 = date2.isoformat()
                    dateV1 = date3.replace("-","")
                    if dateV1 == arg1[:-4]:
                        row_curr = "\"" + dateV1 + "\",\"" + str(arr[row][3]) + "\",\"" + str(arr[row][1]) + "\",\"" + get_student_info(arr[row][1]) + "\"\n"
                        utility.write_to_csv(path_to_out, row_curr)
                else:
                    if file[-4:-1] == "ymd":
                        dateV2 = arr[row][2][:-9]
                        dateV2 = dateV2.replace("-", "")
                        if dateV2 == arg1[:-4]:
                            row_curr = "\"" + dateV2 + "\",\"" + str(arr[row][3]) + "\",\"" + str(arr[row][1]) + "\",\"" + get_student_info(arr[row][1]) + "\"\n"
                            utility.write_to_csv(path_to_out, row_curr)
                    else:
                        if file[-4:-1] == "dmy":
                            dateV3 = arr[row][2][-4:]+""+arr[row][2][-7:-5]+""+arr[row][2][-10:-8]
                            if dateV3 == arg1[:-4]:
                                row_curr = "\"" + dateV3 + "\",\"" + str(arr[row][3]) + "\",\"" + str(arr[row][1]) + "\",\"" + get_student_info(arr[row][1]) + "\"\n"
                                utility.write_to_csv(path_to_out, row_curr)
                #    print(str)
            except:
                l = 0
                # print("Error", arr[row][2])


def write_csv(path, csv_row): # Write output files from Source_1
    full_path = output_folder + path
    with open(full_path, 'a', newline='', encoding="utf-8") as f:
        f.write(csv_row)

# Create distinct files based on date from the external data source
def create_date_dist (argument1):
    print(argument1)
    wb_obj = openpyxl.load_workbook(data_folder+argument1)
    sheet = wb_obj.active
    date = ""
    row_curr = ""
    for row in sheet.iter_rows(max_row=sheet.max_row):
        if row[1].value != "Tarix":
            date = str(row[1].value)
            date = date.split(' ')[0]
            date = datetime.strptime(date, '%Y-%m-%d').strftime('%Y%m%d')
            row_curr = "\""+str(date) +"\",\""+ str(row[3].value) +"\",\""+ str(row[4].value) +"\",\""+ str(row[6].value) +"\",\""+ str(row[8].value) + "\"\n"
            print(row_curr)
            write_csv(date+'.csv', row_curr)
            date = ""
            row_curr = ""


def create_queue(path, q):
    filenames_s = next(walk(path), (None, None, []))[2]
    for file in filenames_s:
        q.put(file, True, None)

def create_list(path, l):
    filenames_s = next(walk(path), (None, None, []))[2]
    for file in filenames_s:
        l.append(file[:-4])

def main():
    quit_flag = False
    #multiprocess addition
    # q = Queue()
    file_list = list()
    procs = []
    # utility.csv2queue(input_file, q)
    # create_queue(data_folder, q)
    # create_queue_infile(data_folder, q)
    # create_list(infolist_folder, file_list)
    # file_list.sort()
    # pool = multiprocessing.Pool(processes=(multiprocessing.cpu_count()-1))
    # while not (q.empty()):
      # proc_f_1 = pool.map_async(search_student, q)   # OLD
      #  proc_f_1 = Process(target=read_xlsx, args=(path_src, q.get())) # OLD
    #   res = pool.apply_async(utility.compare2diff, args=(q.get(), 5, input_file2, 3, output_file, ))

        # complete the processes
    # pool.close()
    # pool.join()

    while(quit_flag == False):
        root = Tk()
        filename = filedialog.askopenfilename(title="Select an xlsx file to Convert to csv",filetypes=(("Xlsx files", "*.xlsx"),))
        dirname = filedialog.askdirectory(title="Select a directory for csv output file")
        root.destroy()
        key_input = input("Select an operation\n1)Convert xlsx file to csv\n2)Compare two lists\n99)Quit\n")
        if key_input == '1':
            start = time.time()
            utility.xlsx2csv(filename, dirname)
            print("Execution time : ", time.time() - start)
        elif key_input == '2':
            start = time.time()
            utility.compare2diff(input_file, input_file2, output_file)
            print("Execution time : ", time.time() - start)
        elif key_input == '99':
            quit_flag = True


if __name__ == '__main__':
    main()


